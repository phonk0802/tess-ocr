import numpy as np
import cv2
import pandas as pd
from PIL import Image
import extract
import pytesseract
from xlsxwriter import Workbook
import os
import openpyxl
from openpyxl.styles import PatternFill

pytesseract.pytesseract.tesseract_cmd = "C:\\Program Files\\Tesseract-OCR\\tesseract.exe"

def binarization(image):
    thresh, im_bw = cv2.threshold(image, 178, 243, cv2.THRESH_BINARY)
    return im_bw

dict_scr_sm1 = [" I: ", " I;", "J;", " I ", " L ", " I-", " kì : ", "kìL"]
dict_scr_sm2 = [" II: ", " II ", " JI ", " II; ", " L :", " JI: ", " JL ", " 1L ", " lI: "]

headers = ["STT", "Họ và tên", "Mã sinh viên", "Ngày sinh", "Khóa", "Khoa", "Ngành", "Điểm TB HKI", "Điểm TB HKII", "Điểm TB năm học"] # List object calls by index, but the dict object calls items randomly
wb = Workbook("DSSV_đọc.xlsx")
wb.encoding = 'utf-8'
ws = wb.add_worksheet("New Sheet")

students = []
files = os.listdir("sinhvien")
for file in files:
    img_file = "sinhvien/" + file
    image = Image.open(img_file)
    image.save(img_file, dpi=(300, 300))
    image = cv2.imread(img_file)
    image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    image = binarization(image)
    data = pytesseract.image_to_string(image, lang="vie")  # kết quả đọc được từ ảnh
    name = ''
    student_id = ''
    date = ''
    course = ''
    department = ''
    major = ''
    score_sm1 = ''  # score semester I
    score_sm2 = ''  # score semester II
    score_avg = ''  # avg score per semester

    lines = data.splitlines()
    for line in lines:
        if score_sm1 == '':
            for i in dict_scr_sm1:
                if i in line:
                    score_sm1 = extract.find_score(line, i)
        if score_sm2 == '':
            for j in dict_scr_sm2:
                if j in line:
                    score_sm2 = extract.find_score(line, j)
        if score_avg == '':
            score_avg = extract.find_score(line, "học:")
        if name == '':
            name = extract.find_name(line)
        if department == '':
            department = extract.find_department(line)
        if major == '':
            major = extract.find_major(line)

    student_id = extract.find_studentID(data)
    course = extract.find_course(data)
    date = extract.find_date(data)

    student = {}
    student["Họ và tên"] = name
    student["Mã sinh viên"] = student_id
    student["Ngày sinh"] = date
    student["Khóa"] = course
    student["Khoa"] = department
    student["Ngành"] = major
    student["Điểm TB HKI"] = score_sm1
    student["Điểm TB HKII"] = score_sm2
    student["Điểm TB năm học"] = score_avg
    students.append(student)

    first_row = 0
    for header in headers:
        col = headers.index(header)
        ws.write(first_row, col, header)

    row = 1
    for student in students:
        for key, value in student.items():
            col = headers.index(key)
            ws.write(row, 0, row)
            ws.write(row, col, value)
        row += 1
wb.close()

df1 = pd.read_excel('DSSV_chuẩn.xlsx')
df2 = pd.read_excel('DSSV_đọc.xlsx')

comparison_values = df1.values == df2.values

rows, cols = np.where(comparison_values == False)
count = 0
count_score = 0
for item in zip(rows,cols):
    df1.iloc[item[0], item[1]] = '{} --> {}'.format(df1.iloc[item[0], item[1]], df2.iloc[item[0], item[1]])
    count += 1

df1.to_excel('Diff.xlsx', index=False, header=True)

wb_diff = openpyxl.load_workbook("Diff.xlsx")
wb_diff.encoding = 'utf-8'
ws_diff = wb_diff.active
fill_cell1 = PatternFill(patternType='solid',
                           fgColor='DA9694')

for row in ws_diff.iter_rows(min_row=1, min_col=1, max_row=31, max_col=10):
    for cell in row:
        if "-->" in str(cell.value):
            ws_diff.cell(row=cell.row, column=cell.column).fill = fill_cell1
            if "-->" in str(cell.value) and cell.column == 8 or cell.column == 9 or cell.column == 10:
                count_score += 1

wb_diff.save('Diff.xlsx')

print("ĐỘ CHÍNH XÁC = " + str((1 - count/300) * 100) + "%")
print("Số đầu điểm đọc chính xác: " + str(90 - count_score) + "/90")